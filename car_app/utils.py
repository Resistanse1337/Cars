import os
import csv
import datetime
from time import time
from random import randint
from pathlib import Path
import openpyxl

from django.core.validators import MaxValueValidator


def current_year():
    return datetime.date.today().year


def max_value_current_year(value):
    return MaxValueValidator(current_year())(value)


def parse_csv(csv_file):
    data = []

    with open(csv_file) as f:
        reader = csv.DictReader(f)

        for row in reader:
            data.append(row)

    return data


def parse_xlsx(xlsx_file):
    data = []

    wb = openpyxl.load_workbook(xlsx_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        rows = ws.iter_rows()
        headers = {i: h.value for i, h in enumerate(next(rows))}

        for i, row in enumerate(rows):
            tmp = {}
            for j, cell in enumerate(row):
                tmp.update({headers[j]: cell.value})

            tmp["sts_date"] = tmp["sts_date"].date()
            data.append(tmp)

    return data


def make_csv(headers, iterable):
    path = os.path.join(Path(__file__).parent, "tmp", f'{int(time())}_{randint(0, 1000)}.csv')

    with open(path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile, delimiter=';',
                                quotechar='|', quoting=csv.QUOTE_MINIMAL)

        writer.writerow(headers)

        for row in iterable:
            writer.writerow(row)

    return path


def make_xlsx(headers, iterable):
    path = os.path.join(Path(__file__).parent, "tmp", f'{int(time())}_{randint(0, 1000)}.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active

    for j, h in enumerate(headers):
        ws.cell(1, j+1, h)

    for i, row in enumerate(iterable):
        current_row = ws.max_row + 1
        for j, cell in enumerate(row):
            ws.cell(current_row, j+1, cell)

    wb.save(path)

    return path


if __name__ == "__main__":
    print(parse_xlsx("/home/resistanse/Projects/cars_service/car_app/tmp/book1.xlsx"))